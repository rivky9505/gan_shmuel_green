from flask import Flask, request, jsonify, Response ,send_from_directory
import json
import mysql.connector
from flask_cors import CORS, cross_origin
import logging
import csv
from openpyxl import Workbook
import xlsxwriter

app = Flask(__name__)

logging.basicConfig(filename='app.log', filemode='w', format='%(name)s - %(levelname)s - %(message)s')

def getMysqlConnection():
    return mysql.connector.connect(user='root', host='mysql', port='3306', password='123', database='billdb')

@app.route("/")
def hello():
    return "Gan Shmuel"

@cross_origin() # Allow all origins all methods.


# GET /health
# - By default returns "OK" and status 200 OK
# -If system depends on external resources (e.g. db), 
# and they are not available (e.g. "select 1;" fails ) 
# then it should return "Failure" and 500 Internal Server Error

@app.route('/health', methods=['GET'])
def get_health():
    db = getMysqlConnection()
    try:
        logging.info('info health test') # CHANGE TO PROPER MESSAGE
        logging.error('error health test') # CHANGE TO PROPER MESSAGE
        sqlstr = "SELECT 1"
        cur = db.cursor()
        cur.execute(sqlstr)
        output_json = cur.fetchall()
    except Exception :
        logging.error('error') # CHANGE TO PROPER MESSAGE
        return jsonify("500 Internal server error")
    finally:
        logging.info("200 OK")
        db.close()
    return jsonify(results=output_json)


def json_to_excel(ws, data, row=0, col=0):
    ws.write('A1', 'Product')
    ws.write('B1', 'Rate')
    ws.write('C1', 'Scope')
    row += 1
    for product_id, rate, scope in data:
        ws.write(row, col, product_id)
        ws.write(row, col + 1, rate)
        ws.write(row, col + 2, scope)
        row += 1

# GET /rates
# Will download a copy of the same excel that was uploaded using POST /rates 
@app.route('/rates', methods=['GET'])
def get_rates():
    db = getMysqlConnection()
    try:
        sqlstr = "SELECT * FROM Rates"
        cur = db.cursor()
        cur.execute(sqlstr)
        output_jason = cur.fetchall()
    except Exception :
        logging.error("ERROR , whilr trying: %s", (sqlstr))
        return jsonify("ERROR , whilr trying: %s", (sqlstr))
    finally:
        logging.info("200 OK SQL completed query: %s", (sqlstr))
        db.close()
    data = output_jason
    wb = xlsxwriter.Workbook("output_from_Rates_Table.xlsx")
    ws = wb.add_worksheet()
    json_to_excel(ws, data)
    wb.close()
    return "Excel Created"

# POST /provider
# Creates a new provider record:
# - name - provider name. must be unique.
# Returns a unique provider id as json: { "id":<str>}
@cross_origin() # allow all origins all methods.

@app.route('/selectAll', methods=['GET'])
def selectAll():
    db = getMysqlConnection()
    try:
        data_query = "SELECT * from Provider"
        logging.info("This is an select all request massege")
        cur = db.cursor()
        cur.execute(data_query)
        
    except Exception as e:
        return("Error in SQL:\n", e)
    finally:
        output_json = cur.fetchall()
        db.close()
        return jsonify(results=output_json)
        # return "Hello"
@cross_origin() # allow all origins all methods.

@app.route('/provider/<provider_name>', methods=['GET','POST'])
def insert_provider(provider_name):
    db = getMysqlConnection()
    try:
        data_query = "INSERT INTO Provider (`name`) VALUES  (%s)"
        data=(provider_name,)
        logging.info("This is an select all request massege")
        cur = db.cursor()  
        cur.execute(data_query,data)
        output_json = cur.lastrowid
        output_json = cur.fetchone()
    except Exception as e:
        print('Could not save duplicate name', str(e))
    finally:
        db.close()
        return jsonify( "id:",(output_json))

#PUT /provider/{id} can be used to update provider name 
@app.route('/provider/<id>', methods=['PUT'])
def putprovider(id):
    try:
        newname = request.form["newname"]
        #return newname
        db = getMysqlConnection()
        cur = db.cursor()  
        cur.execute('UPDATE Provider SET name = ' + '"' +str(newname)+ '"' + ' WHERE id =' + id)
        db.commit()
        cur.close()
        db.close()
        logging.info('info') # CHANGE TO PROPER MESSAGE
        return id
    except Exception as e:
        logging.error('error') # CHANGE TO PROPER MESSAGE
        return str(e)
    



# POST /rates
# - file=<filename>
# Will upload new rates from an excel file in "/in" folder. Rate excel has the following columns:
# - Product - a product id
# - Rate - integer (in agorot)
# - Scope - ALL or A provider id. 
# The new rates over-write the old ones
# A scoped rate has higher precedence than an "ALL" rate

@app.route("/rates",methods=["POST"])
def postrates():
    #filename = "./in/rates.xlsx"
    try:
        details = request.form
        filename = str(details["file"])
        db = getMysqlConnection()
        wb = load_workbook(filename)
        ws = wb.get_active_sheet()
        cur = db.cursor()
        cur.execute('TRUNCATE TABLE Rates') 
        query = "INSERT INTO Rates (product_id, rate, scope) VALUES (%s, %s, %s)" #INSERT
        row = 2
        while ws.cell(row, 1).value is not None:
            product = ws.cell(row, 1).value
            rate = ws.cell(row, 2).value
            scope = ws.cell(row, 3).value
            i_tuple = (product, rate, scope)
            cur.execute(query, i_tuple)
            row += 1

        db.commit()
        cur.close()
        db.close()
        logging.info('info') # CHANGE TO PROPER MESSAGE
        return "RATES UPLOADED"
    except Exception as e:
        logging.error('error') # CHANGE TO PROPER MESSAGE
        return e


# POST /truck
# registers a truck in the system
# - provider - known provider id
# - id - the truck license plate 

@app.route('/truck/<provider_id>/<truck_lisence>', methods=['GET','POST'])
def inserttruck(provider_id, truck_lisence):
    try:
        db = getMysqlConnection()
        data_query2="SELECT id FROM Provider WHERE id="+str(provider_id)
        cur = db.cursor()
        cur.execute(data_query2)
        if cur.fetchone() != None:
            data_query = "INSERT  INTO Trucks (`id`,`provider_id`) VALUES  (%s,%s)"
            data=(truck_lisence,provider_id)
            cur.execute(data_query,data)
        db.commit()
        cur.close()
        db.close()
        return jsonify("OK")
    except Exception as e:
        logging.error('error') # CHANGE TO PROPER MESSAGE
        return str(e)
    

# PUT /truck/{id} can be used to update provider id
@app.route('/truck/<id>', methods=["PUT"])
def updatetruck(id):
    try:
        db = getMysqlConnection()
        cur = db.cursor()  
        cur.execute('')
        db.commit()
        cur.close()
        db.close()
        logging.info('info') # CHANGE TO PROPER MESSAGE
        return "OK"
    except Exception as e:
        logging.error('error') # CHANGE TO PROPER MESSAGE
        return str(e)


# GET /truck/<id>?from=t1&to=t2
# - id is the truck license. 404 will be returned if non-existent
# - t1,t2 - date-time stamps, formatted as yyyymmddhhmmss. server time is assumed.
# default t1 is "1st of month at 000000". default t2 is "now". 
# Returns a json:
# { "id": <str>,
#   "tara": <int>, // last known tara in kg
#   "sessions": [ <id1>,...] 
#}
@app.route('/truck/<id>', methods=["GET"])
def truckinfo(id):
    try:
        #return id
        #return id+str(request.args.get('from')+str(request.args.get('to')))
        db = getMysqlConnection()
        cur = db.cursor()  
        cur.execute('SELECT id , provider_id FROM Trucks WHERE id='+'"' + id + '"')
        results = cur.fetchall()
        return str(results)
        #HERE WE SHOULD MAKE A REQUEST TO WEIGHT API AND GET WITH THE ID BETWEEN DATES BY ID ?
        db.commit()
        cur.close()
        db.close()
        logging.info('info') # CHANGE TO PROPER MESSAGE
        tempJson = { "id"}
        return "OK"
    except Exception as e:
        logging.error('error') # CHANGE TO PROPER MESSAGE
        return str(e)


# GET /bill/<id>?from=t1&to=t2
# - id is provider id
# - t1,t2 - date-time stamps, formatted as yyyymmddhhmmss. server time is assumed.
# default t1 is "1st of month at 000000". default t2 is "now". 
# Returns a json:
# {
#   "id": <str>,
#   "name": <str>,
#   "from": <str>,
#   "to": <str>,
#   "truckCount": <int>,
#   "sessionCount": <int>,
#   "products": [
#     { "product":<str>,
#       "count": <str>, // number of sessions
#       "amount": <int>, // total kg
#       "rate": <int>, // agorot
#       "pay": <int> // agorot
#     },...
#   ],
#   "total": <int> // agorot
# }
@app.route('/bill/<id>?from=<t1>&to=<t2>', methods=["GET"])
def getbilling(id):
    try:
        db = getMysqlConnection()
        cur = db.cursor()  
        cur.execute('')
        db.commit()
        cur.close()
        db.close()
        logging.info('info') # CHANGE TO PROPER MESSAGE
        return "OK"
    except Exception as e:
        logging.error('error') # CHANGE TO PROPER MESSAGE
        return str(e)

@app.route('/getlogs', methods=["GET"])
def getlogs():
    try:
        with open('app.log', 'r') as file:
            return file.read()
    except Exception as e:
        logging.error('file not found')
        return str(e)

if __name__ == "__main__":
    app.run(debug=True,host='0.0.0.0')
