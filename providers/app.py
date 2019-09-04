from flask import Flask, request, jsonify, Response ,send_from_directory , render_template ,json
import json
import mysql.connector
from flask_cors import CORS, cross_origin
import logging
import csv
from openpyxl import Workbook , load_workbook
import xlsxwriter
import os.path
import requests
from datetime import datetime
import bill 
app = Flask(__name__)

logging.basicConfig(filename='app.log', filemode='w', format='%(name)s - %(levelname)s - %(message)s')

def getMysqlConnection():
    return mysql.connector.connect(user='root', host='mysql', port='3306', password='root', database='billdb')


@app.route("/")
def hello():
    return render_template('ProviderMainPage.html')

@cross_origin()


#ERROR CODES 
#(0) - SUCCESS
#(-1) - 500 INTERNAL SERVER ERROR
#(-2) - DATABASE CONNECTION ERROR (HTTP error 503 Service Unavailable)
#(-3) - ERROR EXECUTING QUERY IN DATABASE
#(-4) - I/O ERROR
#(-5) - USER ERROR MISSING PARAMETER IN URL QUERY (HTTP error 400 Bad Request Error)


# GET /health
# - By default returns "OK" and status 200 OK
# -If system depends on external resources (e.g. db), 
# and they are not available (e.g. "select 1;" fails ) 
# then it should return "Failure" and 500 Internal Server Error
@app.route('/health', methods=['GET'])
def checkhealth():
    try:
        db = getMysqlConnection()
    except:
        return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 503
    try:
        query = "SELECT 1"
        cur = db.cursor()
        cur.execute(query)
        result = cur.fetchall()
        logging.info('[GET][SUCCESS] health request . QUERY:' + query)
        return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK" }) , 200 
    except Exception as e:
        logging.error('[GET][FAILURE] /health request . QUERY:' + query)
        return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500 
    finally:
        logging.info("200 OK")
        db.close()


def json_to_excel(ws, data, row=0, col=0):
    ws.write('A1', 'Product')
    ws.write('B1', 'Rate')
    ws.write('C1', 'Scope')
    row += 1
    for product_id, rate, scope in data:
        ws.write(row, col, str(product_id))
        ws.write(row, col + 1, str(rate))
        ws.write(row, col + 2, str(scope))
        row += 1

# GET /rates
# Will download a copy of the same excel that was uploaded using POST /rates 
@app.route('/rates', methods=['GET'])
def get_rates():
    try:
        db = getMysqlConnection()
    except:
        return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 503
    try:
        sqlstr = "SELECT * FROM Rates"
        cur = db.cursor()
        cur.execute(sqlstr)
        output_jason = cur.fetchall()
        db.close()
        logging.info("[GET][SUCCESS] rates request - : %s", (sqlstr))
    except Exception :
        logging.error("[GET][FAILURE] rates request , ON QUERY: %s", (sqlstr))
        return jsonify({ "errorCode" : -3 , "errorDescription" : "ERROR EXECUTING QUERY IN DATABASE" }) , 500
    try:  # Create and save Excel file
        dir_name = "out"
        file_name = "output.xlsx"
        excel_path = "./" + dir_name + "/" + file_name
        data = output_jason
        wb = xlsxwriter.Workbook(excel_path)
        ws = wb.add_worksheet()
        json_to_excel(ws, data)
        wb.close()
        logging.info("[GET][SUCCESS] rates request : Excel file created in: %s", (excel_path))
    except:
        logging.error("[GET][FAILURE] rates request : Excel file NOT created in: %s", (excel_path))
        return jsonify({ "errorCode" : -4 , "errorDescription" : "I/O ERROR : writing Excel file" }) , 500
    try: # send excel file as http response
        if os.path.exists(excel_path):
            logging.info("[GET][SUCCESS] rates request : Excel file from: %s was sent for download", (excel_path))
            return send_from_directory(dir_name, filename=file_name, as_attachment=True, attachment_filename="Rates.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except:
        logging.error("[GET][FAILURE] rates request : Excel file from: %s was NOT sent for download", (excel_path))
        return jsonify({ "errorCode" : -1 , "errorDescription" : "status 404 Not Found : Excel file not found" }) , 500


        

# POST /provider
# Creates a new provider record:
# - name - provider name. must be unique.
# Returns a unique provider id as json: { "id":<str>}

@app.route('/provider/<provider_name>', methods=['POST'])
def insert_provider(provider_name):
    try:
        db = getMysqlConnection()
    except:
        return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 503

    try:
        query_string = "INSERT INTO Provider (name) "
        query_string += "SELECT * FROM (SELECT '" + provider_name + "') AS tmp "
        query_string += "WHERE NOT EXISTS ("
        query_string += "SELECT name FROM Provider WHERE name = '" + provider_name + "'"
        query_string += ") LIMIT 1;"
        cur = db.cursor()  
        cur.execute(query_string)
        db.close()
        logging.info("[POST][SUCCESS] provider/%s", (provider_name,))
        return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK"  , "result": "MYSQL query completed"}) , 200
    except Exception as e:
        logging.info('[POST][FAILURE] while trying:', str(e))
        return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500   

        
#PUT /provider/{id} can be used to update provider name 
# @app.route('/provider/<id>', methods=['PUT'])
# def putprovider(id):
#     try:
#         newname = request.form["newname"]
#         #return newname
#         db = getMysqlConnection()
#         cur = db.cursor()  
#         cur.execute('UPDATE Provider SET name = ' + '"' +str(newname)+ '"' + ' WHERE id =' + id)
#         db.commit()
#         cur.close()
#         db.close()
#         logging.info('[PUT][SUCCESS] provider/<id>') # CHANGE TO PROPER MESSAGE
#         return id
#     except Exception as e:
#         logging.error('[PUT][FAILURE] provider/<id>') # CHANGE TO PROPER MESSAGE
#         return str(e)


# @app.route('/provider/<id>', methods=['PUT'])
# def putprovider2(id):
#     try:
#         db = getMysqlConnection()
#     except:
#         return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 200
    
#     try:
#         newname = request.form["newname"]
#     except:
#         return jsonify({ "errorCode" : -5 , "errorDescription" : "ERROR NO PARAMETERS PASSED" }) , 200
        
#     try:
#         cur = db.cursor()  
#         cur.execute('UPDATE Provider SET name = ' + '"' +str(newname)+ '"' + ' WHERE id =' + id)
#         db.commit()
#         cur.close()
#         db.close()
#         logging.info('[PUT][SUCCESS] provider/<id>') 
#         return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK" }) , 200
#     except Exception as e:
#         logging.error('[PUT][FAILURE] provider/<id>') 
#         return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500




@app.route('/provider', methods=['PUT'])
def putprovider22():
    try:
        db = getMysqlConnection()
    except:
        return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 503
    try:
        json = request.get_json()
        id = str(json["id"])
        newname = str(json["newname"])
        #newname = request.form["newname"]
    except:
        return jsonify({ "errorCode" : -5 , "errorDescription" : "ERROR/WRONG NO PARAMETERS PASSED" }) , 400
    try:    
        cur = db.cursor()  
        cur.execute('UPDATE Provider SET name = ' + '"' +str(newname)+ '"' + ' WHERE id =' + id)
    except:
        return jsonify({ "errorCode" : -3 , "errorDescription" : "ERROR EXECUTING QUERY IN DATABASE" }) , 500
    try:
        db.commit()
        cur.close()
        db.close()
        logging.info('[PUT][SUCCESS] provider/<id>') 
        return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK" }) , 200
    except Exception as e:
        logging.error('[PUT][FAILURE] provider/<id>') 
        return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500


# POST /rates
# - file=<filename>
# Will upload new rates from an excel file in "/in" folder. Rate excel has the following columns:
# - Product - a product id
# - Rate - integer (in agorot)
# - Scope - ALL or A provider id. 
# The new rates over-write the old ones
# A scoped rate has higher precedence than an "ALL" rate

# @app.route("/rates",methods=["POST"])
# def postrates():
#     #filename = "./in/rates.xlsx"
    
#     try:
#         details = request.form
#         filename = str(details["file"])
        
#         db = getMysqlConnection()
#         #wb = load_workbook(filename)
        
#         wb = load_workbook('./in/'+filename)
        
#         ws = wb.get_active_sheet()
#         cur = db.cursor()
        
#         cur.execute('TRUNCATE TABLE Rates') 
#         query = "INSERT INTO Rates (product_id, rate, scope) VALUES (%s, %s, %s)" #INSERT
#         row = 2
#         while ws.cell(row, 1).value is not None:
#             product = ws.cell(row, 1).value
#             rate = ws.cell(row, 2).value
#             scope = ws.cell(row, 3).value
#             i_tuple = (product, rate, scope)
#             cur.execute(query, i_tuple)
#             row += 1

#         db.commit()
#         cur.close()
#         db.close()
#         logging.info('[POST][SUCCESS] /rates ') # CHANGE TO PROPER MESSAGE
#         return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK" }) , 200
#     except Exception as e:
#         logging.error('[POST][FAILURE] /rates') # CHANGE TO PROPER MESSAGE
#         return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500


@app.route("/rates",methods=["POST"])
def postrates():
    #filename = "./in/rates.xlsx"
    try:
        db = getMysqlConnection()
    except:
        return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 503
    try:
        filename_tmp = request.get_json()
        filename = str(filename_tmp["file"])
    except:
        return jsonify({ "errorCode" : -5 , "errorDescription" : "NO PARAMETERS PASSED" }) , 500
        
    try:
        wb = load_workbook("./in/" + filename) # rates2.xlsx")
    except: 
        return jsonify({ "errorCode" : -4 , "errorDescription" : "FILE NOT FOUND" }) , 500
    try:
        ws = wb.get_active_sheet()
        cur = db.cursor()
        cur.execute('TRUNCATE TABLE Rates') 
        query = "INSERT INTO Rates (product_id, rate, scope) VALUES (%s, %s, %s)" #INSERT
    except:
        return jsonify({ "errorCode" : -5 , "errorDescription" : "DB ERROR WRONG PARAMETERS PASSED" }) , 500
        
    try:    
        row = 2
        while ws.cell(row, 1).value is not None:
            product = ws.cell(row, 1).value
            rate = ws.cell(row, 2).value
            scope = ws.cell(row, 3).value
            i_tuple = (product, rate, scope)
            cur.execute(query, i_tuple)
            row += 1
        db.commit()
        cur.close()
        db.close()
        logging.info('[POST][SUCCESS] /rates ') # CHANGE TO PROPER MESSAGE
        return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK" }) , 200
    except Exception as e:
        logging.error('[POST][FAILURE] /rates') # CHANGE TO PROPER MESSAGE
        return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500

# POST /truck
# registers a truck in the system
# - provider - known provider id
# - id - the truck license plate
# This request needs two argumnets.
# Implenting as a query string in url
# http://localhost:5000/truck?id=222-33-111&name=new_provider_for_truck

@app.route('/truck', methods=['POST'])
def inserttruck():
    # get values from query string
    result_message = ""
    result_count_string = ""
    truck_id = ""
    provider_name = ""
    if request.args.get('id') != None:
        truck_id = request.args.get('id')
    else:
        logging.error('[POST][FAILURE] /truck : USER ERROR : MISSING PARAMETER IN URL QUERY')
        return jsonify({ "errorCode" : -5 , "errorDescription" : "USER ERROR MISSING PARAMETER IN URL QUERY" }) , 400

    if request.args.get('name'):
        provider_name = request.args.get('name')
    else:
        logging.error('[POST][FAILURE] /truck : USER ERROR : MISSING PARAMETER IN URL QUERY')
        return jsonify({ "errorCode" : -5 , "errorDescription" : "USER ERROR MISSING PARAMETER IN URL QUERY" }), 400
    
    try:
        db = getMysqlConnection()
    except:
        return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 503
    
    try:
        cur = db.cursor()
        # get id of provider (owner of the truck id)
        querystr = "SELECT id FROM Provider WHERE name = '" + provider_name + "'"
        cur.execute(querystr)
        query_result = cur.fetchall()
        result_count_string = "   Result count: " + str(cur.rowcount)
        if cur.rowcount > 0: # test if there is at least one record
            provider_id = str(query_result[0][0])
            # count how many records have the desired truck id
            querystr = "SELECT COUNT(IF(id='" + truck_id + "',1, NULL)) 'id' FROM Trucks"
            cur.execute(querystr)
            query_result = cur.fetchall()
            if int(query_result[0][0]) > 0: # if more than 0, then don't create the new record.
                result_message = "[POST][FAILURE] /truck : Truck no: " + truck_id + " already exists! Cant create new truck record with the same id."
                logging.info(result_message)
                return jsonify({ "errorCode" : -5 , "errorDescription" : "status 200 OK"  , "result": result_message}) , 200 
            else: # Truck is doesn't exsists -> create new record in table
                querystr = "INSERT  INTO Trucks (`id`,`provider_id`) VALUES  ('" + truck_id + "', " + provider_id + ")"
                cur.execute(querystr)
                cur.close()
                db.close()
                result_message = "[POST][SUCCESS] /truck : Updated Truck no: " + truck_id + " for provider: " + provider_name
                logging.info(result_message)
                return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK"  , "result": result_message}) , 200 
        else: # No id of provider (owner of the truck id)
            result_message = "No provider with this name: " + provider_name
            logging.info(result_message)
            return jsonify({ "errorCode" : -5 , "errorDescription" : "status 200 OK"  , "result": result_message}) , 200 
    except Exception as e:
        logging.error('[POST][FAILURE] /truck : QUERY:' + querystr  +" == " + str(e))
        #return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500
        return str(e) + "\n" + querystr

# PUT /truck/{id} can be used to update provider id
# This request needs two argumnets.
# Implenting as a query string in url
# http://localhost:5000/truck?id=222-33-111&name=new_provider_for_truck
@app.route('/truck', methods=["PUT"])
def updatetruck():
    # get values from query string
    result_message = ""
    result_count_string = ""
    truck_id = ""
    provider_name = ""
    if request.args.get('id') != None:
        truck_id = request.args.get('id')
    else:
        logging.error('[PUT][FAILURE] /truck/ : USER ERROR : MISSING PARAMETER IN URL QUERY')
        return jsonify({ "errorCode" : -5 , "errorDescription" : "USER ERROR MISSING PARAMETER IN URL QUERY" }) , 400

    if request.args.get('name'):
        provider_name = request.args.get('name')
    else:
        logging.error('[PUT][FAILURE] /truck/ : USER ERROR : MISSING PARAMETER IN URL QUERY')
        return jsonify({ "errorCode" : -5 , "errorDescription" : "USER ERROR MISSING PARAMETER IN URL QUERY" }) , 400
    
    try:
        db = getMysqlConnection()
    except:
        return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 503
    
    try:
        cur = db.cursor()
        # get id of provider (owner of the truck id)
        querystr = "SELECT id FROM Provider WHERE name = '" + provider_name + "'"
        cur.execute(querystr)
        query_result = cur.fetchall()
        result_count_string = "   Result count: " + str(cur.rowcount)
        if cur.rowcount > 0: # test if there is at least one record
            provider_id = str(query_result[0][0])
            # count how many records have the desired truck id
            querystr = "SELECT COUNT(IF(id='" + truck_id + "',1, NULL)) 'id' FROM Trucks"
            cur.execute(querystr)
            query_result = cur.fetchall()
            if int(query_result[0][0]) > 0: # if more than 0, then update the record.
                querystr = "UPDATE Trucks SET provider_id = '" + provider_id + "' WHERE id = '" + truck_id + "'" 
                cur.execute(querystr)
                db.commit()
                cur.close()
                db.close()
                result_message = "[PUT][SUCCESS] /truck/ : Updated Truck no: " + truck_id + " for provider: " + provider_name
                logging.info(result_message)
                return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK"  , "result": result_message}) , 200 
            else:
                result_message = "No Truck ID with this id: " + truck_id
                logging.info(result_message)
                return jsonify({ "errorCode" : -5 , "errorDescription" : "status 200 OK"  , "result": result_message}) , 200 
        else: # No id of provider (owner of the truck id)
            result_message = "No provider with this name: " + provider_name
            logging.info(result_message)
            return jsonify({ "errorCode" : -5 , "errorDescription" : "status 200 OK"  , "result": result_message}) , 200 
    except Exception as e:
        logging.error('[PUT][FAILURE] /truck/ : QUERY:' + querystr)
        return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500


# GET /truck/<id>?from=t1&to=t2
# - id is the truck license. 404 will be returned if non-existent
# - t1,t2 - date-time stamps, formatted as yyyymmddhhmmss. server time is assumed.
# default t1 is "1st of month at 000000". default t2 is "now". 
# Returns a json:
# { "id": <str>,
#   "tara": <int>, // last known tara in kg
#   "sessions": [ <id1>,...] 
#}
@app.route('/truck/<id>', methods=["GET"]) #?from=t1&to=t2
def truckinfo(id):
    try:
        fromm = str(request.args.get('from'))
        to = str(request.args.get('to'))
        resp = requests.get('http://green.develeap.com:8080/item/'+ id +' ?from='+ fromm +'&to='+ to +'')
        json_content = json.dumps(resp.json())
        return '{ "errorCode" : 0 , "errorDescription" : "status 200 OK" , "data" :' + str(json_content) + ' }' , 200
         
        #return id
        #return id+str(request.args.get('from')+str(request.args.get('to')))
        #db = getMysqlConnection()
        #cur = db.cursor() 
        #cur.execute('SELECT id , provider_id FROM Trucks WHERE id='+'"' + id + '"')
        #results = cur.fetchall()
        #return str(results)
        ##HERE WE SHOULD MAKE A REQUEST TO WEIGHT API AND GET WITH THE ID BETWEEN DATES BY ID ?
        #db.commit()
        #cur.close()
        #db.close()
        #logging.info('[GET][SUCCESS] /truck/<id>') # CHANGE TO PROPER MESSAGE
        #tempJson = { "id"}
        #return "OK"
    except Exception as e:
        logging.error('[GET][FAILURE] /truck/<id>') # CHANGE TO PROPER MESSAGE
        return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500


#FOR TESTING
@app.route('/truck2/<id>', methods=["GET"])
def truckinfo2(id):
    try:
        db = getMysqlConnection()
    except:
        return jsonify({ "errorCode" : -2 , "errorDescription" : "ERROR ESTABLISHING A DATABASE CONNECTION" }) , 503
    try:
        #return id
        #return id+str(request.args.get('from')+str(request.args.get('to')))
        cur = db.cursor()  
        cur.execute('SELECT id , provider_id FROM Trucks WHERE id='+'"' + id + '"')
        results = cur.fetchall()
        return str(results)
        #HERE WE SHOULD MAKE A REQUEST TO WEIGHT API AND GET WITH THE ID BETWEEN DATES BY ID ?
        db.commit()
        cur.close()
        db.close()
        logging.info('[GET][SUCCESS] /truck/<id>') # CHANGE TO PROPER MESSAGE
        tempJson = { "id"}
        return jsonify({ "errorCode" : 0 , "errorDescription" : "status 200 OK" }) , 200
    except Exception as e:
        logging.error('[GET][FAILURE] /truck/<id>') # CHANGE TO PROPER MESSAGE
        return jsonify({ "errorCode" : -1 , "errorDescription" : "500 Internal server error" }) , 500

# GET /bill/<id>?from=t1&to=t2
# - id is provider id
# - t1,t2 - date-time stamps, formatted as yyyymmddhhmmss. server time is assumed.
# default t1 is "1st of month at 000000". default t2 is "now". 
# Returns a json:
# {
#   "id": <str>,
#   "name": <str>,
#   "from": <str>,
#   "to": <str>,
#   "truckCount": <int>,
#   "sessionCount": <int>,
#   "products": [
#     { "product":<str>,
#       "count": <str>, // number of sessions
#       "amount": <int>, // total kg
#       "rate": <int>, // agorot
#       "pay": <int> // agorot
#     },...
#   ],
#   "total": <int> // agorot
# }
@app.route('/bill/<id>', methods=["GET"])
def getbilling(id):
    try:
        # id
        result={"id" : id}
        # name
        name = bill.get_provider_name(id)[0]
        result.update({"name" : name})
        # t1 & t2
        now = datetime.now()
        t1 = now.strftime("%Y%m01000000")
        t2 = now.strftime("%Y%m%d%H%M%S")
        if request.args.get('t1')!=None:
            t1 = request.args.get('t1')
        if request.args.get('t2')!=None:
            t2 = request.args.get('t2')
        result.update({ "from" : t1 })
        result.update({ "to" : t2 })
        trucks_list=bill.find_providers_trucks(id)
        weights_list=bill.get_all_sessions_in_array(t1,t2)
        rates_dictionary=bill.get_rates()
        # sessionCount
        GlobalSessionsCount=0
        # products
        products=[]
        # truck_in_weights
        trucks_in_weights = set()
        # foreach truck - look for its sessions/weights
        for truck in trucks_list:
            truck_number = str(truck[0])
            truck_sessions_count=0
            for weight in weights_list:
                if weight["truck"] == truck_number:
                    truck_sessions_count += 1
                    GlobalSessionsCount += 1
                    trucks_in_weights.add(truck_number)
                    # if product exist in products - update it
                    flag = False
                    for obj in products:
                        if weight["produce"]==obj["product"]:
                            
                            flag = True
                            amount = int(weight["neto"]) + int(obj["amount"])
                            # return str(obj["amount"])
                            count = int(obj["count"]) +1
                            pay = amount * int(obj["rate"])
                            obj.update({ "amount" : amount , "count" : count ,  "pay" : pay})
                    if flag == False:
                        product=dict()
                        rate = ""
                        for obj in rates_dictionary:
                            if obj["product_id"] == weight["produce"]: 
                                if obj["scope"] == id:
                                    rate = obj["rate"]
                                    break
                                elif obj["scope"] == "All" :
                                    rate = obj["rate"]
                        # pay
                        pay = int(weight["neto"]) * rate
                        product = { "product" : weight["produce"] , "count" : 1 , "amount" : weight["neto"] , "rate" :  rate , "pay" : pay  }
                        products.append(product)

        # set total
        total=0
        for obj in products:
            total += int(obj["pay"])
        result.update({ "truckCount" : len(trucks_in_weights) , "sessionCount" : GlobalSessionsCount , "products" : products , "total" : total })
        

        return jsonify(result)        
        db.commit()
        cur.close()
        db.close()
        logging.info('[GET][SUCCESS] /bill/<id>?from=<t1>&to=<t2>') # CHANGE TO PROPER MESSAGE
    except Exception as e:
        logging.error('[GET][FAILURE] /bill/<id>?from=<t1>&to=<t2>') # CHANGE TO PROPER MESSAGE
        return str(e)



@app.route('/getlogs', methods=["GET"])
def getlogs():
    try:
        with open('app.log', 'r') as file:
            return file.read()
    except Exception as e:
        logging.error('file not found')
        return str(e)


if __name__ == "__main__":
    app.run(debug=True,host='0.0.0.0')
